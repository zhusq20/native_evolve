#!/usr/bin/env python3
"""Offline validation of the agentic SB solve (#5) — NO claude spend.

Drives env.agentic_attempt with a FAKE call_claude and asserts the validity-critical invariants:
  1. GOLD ISOLATION: the sandbox the agent runs in contains the INPUT only — never any *golden* file.
  2. NATIVE SKILL INSTALL: named skills land in sandbox/.claude/skills/<name>/SKILL.md (discoverable).
  3. EXTRACTION PRECEDENCE: the graded artifact prefers sandbox/solution.py, else the fenced block.
  4. COST PASSTHROUGH: the per-session cost is returned.
  5. PROMPT: the skill line appears iff a native skill is installed; the INPUT_PATH/solution.py contract holds.
Run: python3 eval/test_agentic.py
"""
import os
import pathlib
import sys
import tempfile

import openpyxl

EVAL = pathlib.Path(__file__).resolve().parent
sys.path.insert(0, str(EVAL))
sys.path.insert(0, str(EVAL / "envs"))
import spreadsheetbench as sb  # noqa: E402

ENGINE = EVAL.parent / "engine"
SKILL_NAME = "self-verify-and-repair"
SKILL = (ENGINE / "skills" / SKILL_NAME / "SKILL.md").read_text(encoding="utf-8")
NATIVE = [(SKILL_NAME, SKILL)]


def _make_task():
    """A minimal real SB-shaped task: a task dir with one (init, golden) pair."""
    d = tempfile.mkdtemp(prefix="sb_task_")
    wb = openpyxl.Workbook(); ws = wb.active; ws["A1"] = "n"; ws["A2"] = 3; ws["A3"] = 4
    wb.save(os.path.join(d, "0_init.xlsx"))
    wg = openpyxl.Workbook(); wsg = wg.active; wsg["A1"] = "n"; wsg["A2"] = 3; wsg["A3"] = 4; wsg["B1"] = 7
    wg.save(os.path.join(d, "0_golden.xlsx"))                       # the SECRET the agent must not see
    return {"id": "t0", "spreadsheet_path": d, "_root": d,
            "instruction": "Put the sum of column A into B1.",
            "instruction_type": "Cell-Level Manipulation", "answer_position": "B1",
            "answer_sheet": "Sheet"}


def _assert_isolation(sandbox):
    assert os.path.exists(os.path.join(sandbox, "input.xlsx")), "input.xlsx missing from sandbox"
    leaked = []
    for root, _dirs, files in os.walk(sandbox):
        for f in files:
            if "golden" in f.lower():
                leaked.append(os.path.join(root, f))
    assert not leaked, "GOLD LEAK into sandbox: %s" % leaked


def test_with_skill():
    task = _make_task()
    seen = {}

    def fake_call(prompt, **kw):
        sandbox = kw["cwd"]
        seen["prompt"] = prompt
        seen["allowed"] = kw.get("allowed_tools")
        seen["perm"] = kw.get("permission_mode")
        seen["max_turns"] = kw.get("max_turns")
        _assert_isolation(sandbox)
        assert os.path.exists(os.path.join(
            sandbox, ".claude", "skills", SKILL_NAME, "SKILL.md")), "skill not installed"
        with open(os.path.join(sandbox, "solution.py"), "w", encoding="utf-8") as f:
            f.write("# SOLUTION_FILE_MARKER\nimport openpyxl\n")
        return ("```python\n# RETURNED_TEXT_MARKER\n```", 0.0012)        # solution.py must WIN over this

    resp, cost = sb.agentic_attempt(task, "", NATIVE, 20, fake_call, want_cost=True)
    assert "SOLUTION_FILE_MARKER" in resp, "did not prefer sandbox/solution.py"
    assert "RETURNED_TEXT_MARKER" not in resp, "fenced text leaked despite solution.py present"
    assert abs(cost - 0.0012) < 1e-9, "cost not passed through: %r" % cost
    assert sb._extract_code(resp).startswith("# SOLUTION_FILE_MARKER"), "extract_code mismatch"
    assert seen["allowed"] == "Read,Write,Edit,Bash,Skill"
    assert seen["perm"] == "bypassPermissions" and seen["max_turns"] == 20
    assert SKILL_NAME in seen["prompt"], "skill line missing when skill present"
    print("ok  test_with_skill (isolation, install, precedence, cost, flags, prompt)")


def test_no_skill_text_fallback():
    task = _make_task()

    def fake_call(prompt, **kw):
        _assert_isolation(kw["cwd"])
        assert not os.path.isdir(os.path.join(kw["cwd"], ".claude", "skills")), "skill dir present in bare arm"
        return ("here is my answer\n```python\n# FROM_TEXT_MARKER\n```", 0.002)   # no solution.py written

    resp = sb.agentic_attempt(task, "", [], 12, fake_call, want_cost=False)
    assert "FROM_TEXT_MARKER" in resp, "did not fall back to fenced block when no solution.py"
    p = sb.build_agentic_prompt(task, "", [])
    assert SKILL_NAME not in p, "skill line present in bare-agentic prompt"
    assert "solution.py" in p and "INPUT_PATH" in p, "prompt missing the solution.py/INPUT_PATH contract"
    # the verify METHODOLOGY must NOT be in the bare prompt (it lives in the skill -> clean ablation)
    assert "reload" not in p.lower() and "verify" not in p.lower(), "verify methodology leaked into bare prompt"
    print("ok  test_no_skill_text_fallback (bare arm: no skill dir, text fallback, prompt clean)")


def test_empty_response():
    task = _make_task()
    resp, cost = sb.agentic_attempt(task, "", [], 5, lambda p, **kw: ("", 0.0), want_cost=True)
    assert resp == "", "empty model output should yield empty resp (score -> miss), got %r" % resp
    ev = sb.score(task, resp)
    assert ev["em"] == 0.0, "empty resp should score a miss"
    print("ok  test_empty_response (graceful empty -> miss)")


if __name__ == "__main__":
    test_with_skill()
    test_no_skill_text_fallback()
    test_empty_response()
    print("\nALL AGENTIC OFFLINE TESTS PASSED")
