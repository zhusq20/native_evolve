"""SpreadsheetBench environment (code-generation over .xlsx, cell-exact scoring).

The thesis-fit env: mid-range accuracy + many DISTINCT reusable procedural skills
(openpyxl idioms, header handling, preserving untouched cells, range math). LLM is
our `claude` CLI; the deterministic harness (code execution + official cell compare)
is reused from SkillOpt's standalone executor/evaluator (openpyxl-only), copied into
eval/envs/sb_lib/ so this project is self-contained.

Task file = the extracted dataset.json; spreadsheet files resolve relative to it.
"""
import glob
import json
import os
import pathlib
import shutil
import sys
import tempfile

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent / "sb_lib"))
import executor as _exec      # noqa: E402  run_generated_code(code, in, out, timeout)
import evaluator as _evalmod  # noqa: E402  evaluate(pred, gold, itype, answer_position)

NAME = "spreadsheetbench"

_SYSTEM = (
    "You are an expert Python programmer specializing in spreadsheet manipulation. "
    "Write a single self-contained Python script that reads the workbook at the path in "
    "variable INPUT_PATH, performs the requested manipulation, and saves the result to "
    "OUTPUT_PATH. Use ONLY the Python standard library and openpyxl (pandas is NOT "
    "available). Do not print. Do not call input(). Do not hardcode file paths or row "
    "counts — iterate over all actual rows/columns. Preserve every other cell unchanged. "
    "Return ONLY the Python code inside a single ```python ... ``` fenced block."
)


def load_tasks(path):
    p = pathlib.Path(path)
    data = json.loads(p.read_text(encoding="utf-8"))
    root = p.parent
    out = []
    for r in data:
        r = dict(r)
        r["question"] = r.get("instruction", "")   # generic key used by the runner
        r["_root"] = str(root)
        out.append(r)
    return out


def _task_dir(task):
    sp = task.get("spreadsheet_path", "spreadsheet/%s" % task.get("id"))
    return sp if os.path.isabs(sp) else os.path.join(task["_root"], sp)


def _test_cases(task_dir):
    """Return [(init_xlsx, golden_xlsx), ...] for a task (verified_400 layout)."""
    pairs = []
    for ip in sorted(glob.glob(os.path.join(task_dir, "*_init.xlsx"))):
        gp = ip.replace("_init.xlsx", "_golden.xlsx")
        if os.path.exists(gp):
            pairs.append((ip, gp))
    if not pairs:  # bare fallback
        bi, bg = os.path.join(task_dir, "initial.xlsx"), os.path.join(task_dir, "golden.xlsx")
        if os.path.exists(bi) and os.path.exists(bg):
            pairs.append((bi, bg))
    return pairs


def _answer_position(task):
    ap = task.get("answer_position", "")
    sheet = task.get("answer_sheet", "")
    if ap and sheet and "!" not in ap:
        return "%s!%s" % (sheet, ap)
    return ap


def _inspect_target_cells(xlsx_path, answer_position, max_cells=12):
    """Inspect the cells the grader will read in a produced workbook. `data_only=False`
    so we SEE whether a cell literally holds a formula string ('=...') — openpyxl stores
    that string verbatim and never evaluates it, so the grader reads the formula text (or a
    stale/None cached value) instead of the computed result and marks it wrong. This is the
    single most common SB failure our trace-blind reflector used to invert. Uses the official
    evaluator's range parser so the cell set matches exactly. Reads NO gold (the produced
    workbook only) -> reusable by the reference-free verify() hook."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:  # noqa: BLE001
        return [{"coord": "?", "note": "could not open produced workbook: %s" % e}]
    out = []
    try:
        for scr in (answer_position or "").split(","):
            scr = scr.strip()
            if not scr:
                continue
            if "!" in scr:
                sn, rng = scr.split("!", 1)
                sn = sn.strip().strip("'\"")
            else:
                sn = wb.sheetnames[0]
                rng = scr
            rng = rng.strip().strip("'\"")
            if sn not in wb.sheetnames:
                out.append({"coord": scr, "note": "sheet %r missing in output" % sn})
                continue
            ws = wb[sn]
            for cn in _evalmod._generate_cell_names(rng):
                v = ws[cn].value
                out.append({
                    "coord": "%s!%s" % (sn, cn),
                    "value": repr(v)[:60],
                    "is_formula_string": isinstance(v, str) and v.startswith("="),
                    "is_none": v is None,
                })
                if len(out) >= max_cells:
                    return out
    finally:
        wb.close()
    return out


def _norm(v):
    """Light value normalization for the semantic diff's mismatch flag (the grader has ALREADY decided
    the task is wrong; this only selects WHICH cells to surface to the reflector). Numbers (and numeric
    strings) compared with float tolerance; other text stripped + casefolded; None stays None."""
    if isinstance(v, bool) or v is None:
        return v
    if isinstance(v, (int, float)):
        return round(float(v), 6)
    s = str(v).strip()
    try:
        return round(float(s), 6)
    except (ValueError, TypeError):
        return s.casefold()


def _gold_vs_pred(pred_path, golden_path, answer_position, max_cells=12):
    """GOLD-GROUNDED semantic diff for trace-grounded reflection: compare the COMPUTED VALUES the grader
    reads in the produced vs the golden workbook, cell by cell. This is the SEMANTIC layer the
    reference-free form check (verify/try_run) is blind to — code that ran, wrote literals into the right
    cells, but computed the WRONG NUMBER. It READS the *_golden* file, so it is for REFLECTION/training
    ONLY: it rides on score() (which already uses gold to grade) and is surfaced solely via
    evidence()/_diagnose at train time; it is NEVER reachable from verify()/try_run() (the gold-free
    repair signal). data_only=True -> cached computed values, so a literal cell is compared on its value
    (a formula-STRING cell reads as None here and is left to the form branch). Cells whose GOLD value is
    None are skipped (no expected value to teach from). Returns [{coord, gold, pred, mismatch}, ...]."""
    try:
        wp = openpyxl.load_workbook(pred_path, data_only=True)
        wg = openpyxl.load_workbook(golden_path, data_only=True)
    except Exception as e:  # noqa: BLE001
        return [{"coord": "?", "note": "could not open workbooks for value diff: %s" % e}]
    out = []
    try:
        for scr in (answer_position or "").split(","):
            scr = scr.strip()
            if not scr:
                continue
            if "!" in scr:
                sn, rng = scr.split("!", 1)
                sn = sn.strip().strip("'\"")
            else:
                sn = wg.sheetnames[0]
                rng = scr
            rng = rng.strip().strip("'\"")
            if sn not in wg.sheetnames:
                continue
            gs = wg[sn]
            ps = wp[sn] if sn in wp.sheetnames else None
            for cn in _evalmod._generate_cell_names(rng):
                gv = gs[cn].value
                if gv is None:                       # no expected value to teach from -> skip
                    continue
                pv = ps[cn].value if ps is not None else None
                out.append({
                    "coord": "%s!%s" % (sn, cn),
                    "gold": repr(gv)[:48],
                    "pred": repr(pv)[:48],
                    "mismatch": _norm(gv) != _norm(pv),
                })
                if len(out) >= max_cells:
                    return out
    finally:
        wg.close(); wp.close()
    return out


def _preview(xlsx, max_rows=5, max_cols=20):
    try:
        wb = openpyxl.load_workbook(xlsx, data_only=False)
    except Exception as e:  # noqa: BLE001
        return "(failed to preview: %s)" % e
    chunks = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        chunks.append("## Sheet: %s (dim=%s, max_row=%s, max_col=%s)"
                      % (sn, ws.dimensions, ws.max_row, ws.max_column))
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows),
                                max_col=min(ws.max_column, max_cols), values_only=False):
            cells = []
            for c in row:
                v = c.value
                s = "" if v is None else (str(v)[:37] + "..." if len(str(v)) > 40 else str(v))
                cells.append("%s=%s" % (c.coordinate, s))
            chunks.append(" | ".join(cells))
        if ws.max_row > max_rows:
            chunks.append("... (%d more rows)" % (ws.max_row - max_rows))
        chunks.append("")
    wb.close()
    return "\n".join(chunks)


def _extract_code(text):
    if not text or "```" not in text:
        return (text or "").strip()
    start = text.find("```")
    nl = text.find("\n", start)
    end = text.find("```", nl + 1)
    if nl == -1 or end == -1:
        return text.strip()
    return text[nl + 1:end].strip()


def build_prompt(task, mem):
    cases = _test_cases(_task_dir(task))
    preview = _preview(cases[0][0]) if cases else "(no input workbook found)"
    body = (
        _SYSTEM + "\n\n"
        "# Instruction\n" + task.get("instruction", "") +
        "\nInstruction type: " + str(task.get("instruction_type", "")) +
        "\nExpected answer position: " + _answer_position(task) +
        "\n\n# Input spreadsheet preview\n" + preview +
        "\n\n# Task\nReturn only a ```python``` code block."
    )
    return (mem + "\n\n" + body) if mem else body


def score(task, response):
    code = _extract_code(response)
    cases = _test_cases(_task_dir(task))
    ap = _answer_position(task)
    itype = task.get("instruction_type", "")
    n_pass, reason, diag = 0, "", None
    tmp = tempfile.mkdtemp(prefix="sb_")
    try:
        for i, (init, golden) in enumerate(cases):
            pred = os.path.join(tmp, "%d_pred.xlsx" % i)
            ok_exec, err = _exec.run_generated_code(code, init, pred, timeout=60)
            if not ok_exec:
                reason = reason or ("exec error: " + (err or "")[:240])
                if diag is None:                  # capture the FIRST failure's full diagnostics
                    diag = {"executed": False, "traceback": (err or ""),
                            "cell_reason": "", "target_cells": []}
                continue
            ev = _evalmod.evaluate(pred, golden, itype, ap)
            if ev.get("ok"):
                n_pass += 1
            else:
                reason = reason or ("wrong cells: " + str(ev.get("reason", ""))[:240])
                if diag is None:                  # inspect pred BEFORE the tempdir is removed
                    diag = {"executed": True, "traceback": "",
                            "cell_reason": ev.get("reason", ""),
                            "target_cells": _inspect_target_cells(pred, ap),
                            # gold-grounded value diff (train-only; surfaced via evidence()/_diagnose,
                            # never via the gold-free verify()/try_run()) -> the SEMANTIC layer
                            "value_diff": _gold_vs_pred(pred, golden, ap)}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    em = 1.0 if (cases and n_pass == len(cases)) else 0.0
    soft = (n_pass / len(cases)) if cases else 0.0
    return {"em": em, "f1": soft, "sub_em": em, "predicted_answer": "(python code)",
            "gold_answers": ["cells %s" % ap], "_reason": reason,
            "_npass": n_pass, "_ncases": len(cases), "_code": code, "_diag": diag}


def summarize(task, response, ev):
    return (
        "USER TASK (SpreadsheetBench, %s):\n%s\n\nANSWER POSITION: %s\n\n"
        "WAS CORRECT: %s  (passed %d/%d test cases)\n"
        "FAILURE REASON: %s\n\n"
        "AGENT CODE (truncated):\n%s"
        % (task.get("instruction_type", ""), task.get("instruction", "")[:900],
           _answer_position(task), ev["em"] == 1.0, ev.get("_npass", 0),
           ev.get("_ncases", 0), ev.get("_reason", "") or "(none)",
           (ev.get("_code", "") or "")[:900])
    )


def _diagnose(diag):
    """Turn the structured _diag captured at score time into a grounded, actionable diagnosis.
    Names the formula-string poison explicitly so the reflector learns the CORRECT lesson
    (compute literals) instead of its trace-blind inverse."""
    if not diag:
        return ""
    if not diag.get("executed", True):
        return ("CODE FAILED TO EXECUTE. Full traceback below — fix the exception (do not guess):\n"
                + (diag.get("traceback") or "")[:1800])
    parts = []
    cells = diag.get("target_cells") or []
    formula = [c for c in cells if c.get("is_formula_string")]
    nones = [c for c in cells if c.get("is_none")]
    if formula:
        parts.append("FORMULA-STRING POISON: the code wrote Excel formula strings into target cells "
                     + ", ".join(c["coord"] for c in formula[:6]) +
                     " (e.g. value=" + (formula[0].get("value", "") or "") + "). openpyxl stores the "
                     "string '=...' verbatim and NEVER evaluates it, so the grader reads the formula "
                     "text / a None cached value and marks it WRONG. FIX: compute the concrete value in "
                     "Python and write that literal, not a formula string.")
    if nones:
        parts.append("Target cells left EMPTY (None): " + ", ".join(c["coord"] for c in nones[:6]) +
                     " — the code did not write the expected output cells.")
    if diag.get("cell_reason"):
        parts.append("Grader's first mismatch — " + str(diag["cell_reason"]))
    value_diff = diag.get("value_diff") or []
    mism = [d for d in value_diff if d.get("mismatch")]
    if cells and not formula and not nones and mism:
        # FORM is clean (ran, wrote literals into the right cells) but the VALUES are WRONG -> the
        # SEMANTIC blind spot the reference-free form/execution check cannot see. Teach the LOGIC,
        # not the form (gated on form-clean so it never competes with the formula/None lessons above).
        parts.append(
            "SEMANTIC ERROR (code ran, target cells hold literals, but the COMPUTED VALUES are WRONG — "
            "the failure a form/execution check CANNOT catch). Per-cell expected-vs-got:\n"
            + "\n".join("  %s: expected %s, got %s" % (d["coord"], d["gold"], d["pred"]) for d in mism[:8])
            + "\nThe bug is in the LOGIC, not the openpyxl form: re-derive HOW the answer should be "
            "computed from the instruction (operation / filter / aggregation / rounding / units / which "
            "rows) and record a transferable rule about THIS KIND of computation — not a formatting tip.")
    elif cells and not formula and not nones:
        parts.append("Values written to target cells: "
                     + "; ".join("%s=%s" % (c["coord"], c.get("value")) for c in cells[:6]))
    return "\n".join(parts)


def evidence(task, response, ev):
    correct = ev["em"] == 1.0
    return {
        "outcome": "PASS" if correct else "FAIL",
        "task": "SpreadsheetBench (%s): %s"
                % (task.get("instruction_type", ""), task.get("instruction", "")[:700]),
        "predicted": "passed %d/%d test cases; answer position %s"
                     % (ev.get("_npass", 0), ev.get("_ncases", 0), _answer_position(task)),
        "gold": "the target cells must hold the correct COMPUTED LITERAL values",
        "diagnosis": "" if correct else _diagnose(ev.get("_diag")),
        "code": (ev.get("_code", "") or "")[:1200],
    }


def verify(task, attempt):
    """REFERENCE-FREE check: run the produced code on the INPUT workbook ONLY (never the *_golden*
    file) and inspect the target cells. Catches the dominant gold-independent failures — no code,
    a crash, empty targets, or a FORMULA STRING openpyxl won't evaluate. Returns None when there's
    nothing to check. Reads NO gold, so it is valid during the frozen TEST phase."""
    code = _extract_code(attempt)
    if not code.strip() or "```" not in (attempt or ""):   # contract: a single fenced python block
        return {"ok": False, "signature": "no-code",
                "feedback": "No ```python``` code block was found in your output. Return a single "
                            "self-contained fenced python block."}
    cases = _test_cases(_task_dir(task))
    if not cases:
        return None                                  # no input workbook -> cannot verify
    init = cases[0][0]                               # FIRST case's INPUT only (never golden)
    ap = _answer_position(task)
    tmp = tempfile.mkdtemp(prefix="sbv_")
    try:
        pred = os.path.join(tmp, "verify_pred.xlsx")
        ok_exec, err = _exec.run_generated_code(code, init, pred, timeout=60)
        if not ok_exec:
            return {"ok": False, "signature": "exec-error",
                    "feedback": "Your code failed to execute on the input workbook. Full traceback:\n"
                                + (err or "")[:1200] + "\nFix the exception and try again."}
        cells = _inspect_target_cells(pred, ap)
        formula = [c for c in cells if c.get("is_formula_string")]
        if formula:
            return {"ok": False, "signature": "formula-string-in-target",
                    "feedback": "Your code wrote Excel FORMULA STRINGS into target cells "
                                + ", ".join(c["coord"] for c in formula[:6]) + " (e.g. "
                                + (formula[0].get("value", "") or "") + "). openpyxl stores '=...' as "
                                "literal text and never computes it, so the grader reads the formula "
                                "text and marks it wrong. Compute the value in Python and write the "
                                "literal result instead of a formula string."}
        nones = [c for c in cells if c.get("is_none")]
        if cells and len(nones) == len(cells):
            return {"ok": False, "signature": "empty-target",
                    "feedback": "All target cells (%s) are empty after running your code — it did "
                                "not write the expected output. Write the computed values into the "
                                "answer cells." % ap}
        return {"ok": True, "signature": "", "feedback": ""}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def try_run(task, attempt):
    """GENERIC, dataset-agnostic execution probe for self_verify (no answer_position, no golden).
    Runs the candidate on the task INPUT — a coding agent runs its own code in any runtime — and
    reports (a) a crash, or (b) formula STRINGS the code WROTE, found by diffing input vs output
    (general openpyxl reasoning, not knowledge of which cells are graded). Returns (ran_ok, feedback);
    ran_ok=None when there is no workbook to run against. NEVER opens the *_golden* file."""
    code = _extract_code(attempt)
    if not code.strip() or "```" not in (attempt or ""):
        return (False, "No ```python``` code block found; return one self-contained fenced block.")
    cases = _test_cases(_task_dir(task))
    if not cases:
        return (None, "")
    init = cases[0][0]                                # INPUT only — never the golden
    tmp = tempfile.mkdtemp(prefix="sbtr_")
    try:
        pred = os.path.join(tmp, "out.xlsx")
        ok_exec, err = _exec.run_generated_code(code, init, pred, timeout=60)
        if not ok_exec:
            return (False, "Your code crashed when run on the input:\n" + (err or "")[:900])
        try:                                          # input/output diff: did the code write formula strings?
            wi = openpyxl.load_workbook(init, data_only=False)
            wo = openpyxl.load_workbook(pred, data_only=False)
            wrote_formula = []
            for sn in wo.sheetnames:
                wso = wo[sn]
                wsi = wi[sn] if sn in wi.sheetnames else None
                for row in wso.iter_rows(max_row=min(wso.max_row, 400)):
                    for c in row:
                        v = c.value
                        if isinstance(v, str) and v.startswith("="):
                            iv = wsi[c.coordinate].value if wsi is not None else None
                            if v != iv:               # the code itself wrote this formula string
                                wrote_formula.append("%s!%s" % (sn, c.coordinate))
                    if len(wrote_formula) >= 6:
                        break
                if len(wrote_formula) >= 6:
                    break
            wi.close(); wo.close()
            if wrote_formula:
                return (False, "Your code wrote Excel FORMULA STRINGS into %s. openpyxl stores '=...' "
                        "as literal text and never computes it, so any reader of the file sees the "
                        "formula text (or None), not the value. Compute the values in Python and write "
                        "the literal results." % ", ".join(wrote_formula[:6]))
        except Exception:
            pass
        return (True, "")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
